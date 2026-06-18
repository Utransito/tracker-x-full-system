
from pathlib import Path
from sqlalchemy.orm import Session
import openpyxl
from .models import User, Store

ROLES = [
    ("Administrador General", "admin", "admin123", "admin"),
    ("Usuario Central de Monitoreo", "central", "123456", "central_monitoreo"),
    ("Analista de Hurtos", "hurtos", "123456", "analista_hurtos"),
    ("Analista de Delitos", "delitos", "123456", "analista_delitos"),
]

def seed_users(db: Session):
    if db.query(User).count() == 0:
        for full_name, username, password, role in ROLES:
            db.add(User(full_name=full_name, username=username, password=password, role=role))
        db.commit()

def _to_str(v):
    return "" if v is None else str(v).replace("\n", " ").strip()

def _to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return None

def seed_stores(db: Session):
    if db.query(Store).first():
        return
    excel_path = Path(__file__).resolve().parent.parent / "lista de tias_geo operaciones.xlsx"
    inserted = 0
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None or row[2] is None:
                continue
            values = list(row) + [None] * (18 - len(row))
            db.add(Store(
                central_code=_to_str(values[0]),
                store_code=_to_str(values[2]),
                format=_to_str(values[4]),
                store_name=_to_str(values[3]),
                region=_to_str(values[6]),
                district=_to_str(values[7]),
                zone=_to_str(values[8]),
                lat=_to_float(values[9]),
                lng=_to_float(values[10]),
                province=_to_str(values[11] or values[5]),
                canton=_to_str(values[12]),
                parish="",
                subzone=_to_str(values[13]),
                police_district=_to_str(values[14]),
                circuit=_to_str(values[15]),
                subcircuit=_to_str(values[16]),
                subcircuit_code=_to_str(values[17]),
            ))
            inserted += 1
        db.commit()
    if inserted == 0:
        fallback = [
            dict(store_code="116", central_code="103", format="TIA", store_name="OLMEDO", region="REGION 2", district="NO APLICA", zone="ZONA 12", lat=-2.198406, lng=-79.884537, province="GUAYAS", canton="GUAYAQUIL", parish="", subzone="DMG", police_district="9 DE OCTUBRE", circuit="CHILE", subcircuit="CHILE 3", subcircuit_code="09D03C01S03"),
        ]
        for row in fallback:
            db.add(Store(**row))
        db.commit()
