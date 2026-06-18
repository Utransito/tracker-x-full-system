from datetime import datetime
# V40 hard fix: suspect status defaults and modalidad catalog
from pathlib import Path
from uuid import uuid4
from typing import Optional, List
from types import SimpleNamespace
from collections import Counter
from tempfile import NamedTemporaryFile
import openpyxl
import json

from fastapi import FastAPI, Request, Depends, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func

from .database import Base, engine, get_db, SessionLocal
from .models import User, Store, Case, CaseFile
from .seed import seed_users, seed_stores
from .pdf_utils import generate_case_pdf, generate_case_sheet_pdf, generate_summary_pdf, generate_preinforme_pdf, generate_executive_dashboard_pdf

app = FastAPI(title="GDN - Centro de Análisis y Seguridad Empresarial")
app.add_middleware(SessionMiddleware, secret_key="trackerx-secret-key-change-me")

BASE_DIR = Path(__file__).resolve().parent.parent
UPLOAD_DIR = BASE_DIR / "uploads"
PDF_DIR = BASE_DIR / "generated_pdfs"
BACKUP_DIR = BASE_DIR / "case_backups"
UPLOAD_DIR.mkdir(exist_ok=True)


@app.get("/uploads/{stored_name}")
def serve_uploaded_file(stored_name: str, request: Request, db: Session = Depends(get_db)):
    require_login(request, db)
    file_path = UPLOAD_DIR / stored_name
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404)
    return FileResponse(str(file_path))

PDF_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)


DESEGREGACION_FILE = BASE_DIR / "DESEGREGACION INFRACCIONES.xlsx"

def load_novelty_catalog():
    fallback_types = ["DELITO", "CONTRAVENCIÓN", "RIESGO", "TRANSITO"]
    fallback_matrix = {
        "DELITO": {},
        "CONTRAVENCIÓN": {},
        "RIESGO": {},
        "TRANSITO": {},
    }
    if not DESEGREGACION_FILE.exists():
        return fallback_types, fallback_matrix

    wb = openpyxl.load_workbook(DESEGREGACION_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [str(c.value).strip().upper() if c.value is not None else "" for c in ws[1]]
    try:
        tipo_idx = headers.index("TIPO DE NOVEDAD")
        subtipo_idx = headers.index("SUBTIPO DE NOVEDAD")
        modalidad_idx = headers.index("MODALIDAD")
    except ValueError:
        return fallback_types, fallback_matrix

    ordered_types = []
    matrix = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        tipo = str(row[tipo_idx]).strip() if row[tipo_idx] not in (None, "") else ""
        subtipo = str(row[subtipo_idx]).strip() if row[subtipo_idx] not in (None, "") else ""
        modalidad = str(row[modalidad_idx]).strip() if modalidad_idx < len(row) and row[modalidad_idx] not in (None, "") else ""
        if not tipo or not subtipo:
            continue

        if tipo not in ordered_types:
            ordered_types.append(tipo)
        matrix.setdefault(tipo, {})
        matrix[tipo].setdefault(subtipo, [])
        if modalidad and modalidad.upper() != "NAN" and modalidad not in matrix[tipo][subtipo]:
            matrix[tipo][subtipo].append(modalidad)

    return ordered_types or fallback_types, matrix or fallback_matrix


app.mount("/static", StaticFiles(directory=str(BASE_DIR / "app" / "static")), name="static")
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
templates = Jinja2Templates(directory=str(BASE_DIR / "app" / "templates"))

Base.metadata.create_all(bind=engine)
with SessionLocal() as db:
    seed_users(db)
    seed_stores(db)

PROCESS_OPTIONS = ["HURTO", "DELITO"]
SUBPROCESS_OPTIONS = ["Seguimiento sospechoso", "Registro novedad"]
SECURITY_COMPANIES = ["SECURYTY WORDL", "LAAR SEGURIDAD"]
ALERT_STATES = ["Confirmado", "Descartado"]
DISCARDED_REASONS = ["No procede a la actividad delictual", "Error por manipular el dispositivo", "Simulacro"]
AFFECTATIONS = ["Directa", "Indirecta"]
NOVELTY_LOCATIONS = ["Interior del local", "Exterior del local"]
PLACES = ["Banco del Pacifico", "Banco de Guayaquil", "Wester Union", "Tuti", "Vía Pública", "Lugares de acceso público", "Local comercial", "Residencias", "Tía"]
PLACE_SUBTYPES = ["Cajero", "Tienda", "Banco del barrio", "Farmacia", "Panadería", "Hotel/motel", "Estacionamiento", "Parque", "Vía pública", "Centro comercial", "Casa/departamento", "Conjunto residencial"]
POLICE_UNITS = [
    "Unidad Policial (Desconcentración)",
    "Dirección General de Seguridad Ciudadana y Orden Público",
    "Subzona de Policía",
    "Distrito de Policía",
    "Circuito de Policía",
    "Subcircuito de Policía",
]
REQUESTED_BY = ["Centro de análisis", "Central de monitoreo", "Jefe de local", "Guardia de seguridad", "Operador CCTV", "Otro"]
NOVELTY_TYPES, NOVELTY_MATRIX = load_novelty_catalog()

MODALIDAD_CATALOGO = [
    "OCULTAMIENTO DE MERCADERIA",
    "DESCUIDEROS",
    "CONSUMO NO FACTURADO",
    "ASALTO",
    "HORAMEN",
    "ESCALAMIENTO",
    "TUMBA PUERTAS",
    "INHABILITACIÓN DE SISTEMAS DE SEGURIDAD Y SERVICIOS BASICOS",
    "INCONSISTENCIA EN FACTURAS",
    "SAQUEO",
    "FALSOS FUNCIONARIOS",
    "VACUNA",
    "NO APLICA",
]
SUSPECT_STATUS = ["Aprehendido", "No aprehendido", "Procesado"]
ANTECEDENTES = ["Reincidente", "Sospechoso no reconocido"]
TIPOS_ORGANIZACION = ["Delincuencia común", "Delincuencia organizada", "Delincuente habitual", "Personas en situación de calle", "Personas en estado de embriaguez", "Personas bajo efecto de sustancias psicotrópicas", "Personas con problemas mentales"]
ORG_NAMES = ["Águilas", "AguilasKiller", "Ak47", "Caballeros Oscuros", "ChoneKiller", "Choneros", "Corvicheros", "Cubanos", "Fatales", "Latin Kings", "Lobos", "Los Tiburones", "Mafia18", "R7", "Tiguerones", "No reconocida"]
NATIONALITIES = ["Ecuador", "Perú", "Colombia", "Argentina", "Brasil", "Venezuela", "Panamá", "México", "EEUU", "Canadá", "España", "Italia", "Uruguay", "Paraguay"]
PRODUCT_TYPES = ["Mercadería", "Dinero", "Bienes particulares"]
MERCH_STATES = ["Recuperada", "Sustraída", "Pagada por consumo"]
LEGAL_STATUSES = ["APREHENDIDO", "NO APREHENDIDO", "PROCESADO", "SOBRESEÍDO", "SENTENCIADO", "AUDIENCIA PENDIENTE"]
PRECAUTIONARY_MEASURES = ["Prisión preventiva", "Presentación periódica", "Prohibición de salida del país", "Medidas alternativas", "Ninguna", "En análisis"]


@app.head("/")
@app.get("/healthz")
def health() -> PlainTextResponse:
    return PlainTextResponse("ok")


def current_user(request: Request, db: Session):
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    return db.query(User).filter(User.id == user_id).first()


def require_login(request: Request, db: Session):
    user = current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    return user


def require_admin(request: Request, db: Session):
    user = require_login(request, db)
    if user.role != "admin":
        raise HTTPException(status_code=403)
    return user


def build_case_number(db: Session, base_code: str):
    year = 2026
    count = db.query(func.count(Case.id)).scalar() or 0
    return f"{base_code}{year}{str(count + 1).zfill(5)}"


def save_upload(upload: Optional[UploadFile], file_type: str, case_id: int, db: Session):
    if not upload or not upload.filename:
        return None
    ext = Path(upload.filename).suffix
    safe_name = f"{uuid4().hex}{ext}"
    target = UPLOAD_DIR / safe_name
    with target.open("wb") as f:
        f.write(upload.file.read())
    rec = CaseFile(case_id=case_id, file_type=file_type, original_name=upload.filename, stored_name=safe_name)
    db.add(rec)
    return rec


def to_json(value: str | None):
    if not value:
        return "[]"
    try:
        json.loads(value)
        return value
    except Exception:
        return "[]"


def proceedings_to_storage(summary_text: str | None, items_json: str | None):
    payload = {"summary": summary_text or "", "items": []}
    if items_json:
        try:
            payload["items"] = json.loads(items_json)
        except Exception:
            payload["items"] = []
    return json.dumps(payload, ensure_ascii=False)

def proceedings_from_storage(raw: str | None):
    if not raw:
        return {"summary": "", "items": []}
    try:
        data = json.loads(raw)
        if isinstance(data, dict):
            return {
                "summary": data.get("summary", "") or "",
                "items": data.get("items", []) or [],
            }
    except Exception:
        pass
    return {"summary": raw, "items": []}

def case_snapshot(case: Case):
    return {c.name: getattr(case, c.name) for c in case.__table__.columns}

def save_case_backup(case: Case):
    snap = case_snapshot(case)
    stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    target = BACKUP_DIR / f"{case.case_number}_{stamp}.json"
    with target.open("w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, default=str, indent=2)
    return str(target)

def case_to_form_data(case: Case):
    def safe_list(raw):
        try:
            data = json.loads(raw or "[]")
            return data if isinstance(data, list) else []
        except Exception:
            return []
    payload = {}
    for c in case.__table__.columns:
        value = getattr(case, c.name)
        if isinstance(value, datetime):
            value = value.isoformat()
        payload[c.name] = value
    payload["victims"] = safe_list(case.victims_json)
    payload["suspects"] = safe_list(case.suspects_json)
    payload["merchandise"] = safe_list(case.merchandise_json)
    payload["vehicles"] = safe_list(case.vehicles_json)
    payload["processed"] = safe_list(case.processed_json)
    payload["photo_notes"] = safe_list(case.photo_notes_json)
    payload["has_victims"] = "SI" if payload["victims"] else "NO"
    payload["has_suspects"] = "SI" if payload["suspects"] else "NO"
    payload["has_vehicles"] = "SI" if payload["vehicles"] else "NO"
    payload["has_processed"] = "SI" if payload["processed"] else "NO"
    payload["has_merchandise"] = "SI" if payload["merchandise"] else "NO"
    payload["has_penal_process"] = "SI" if any([case.delegation, case.prosecutor_office, case.complaint_number, case.proceedings]) else "NO"
    payload["penal_process_type"] = "FLAGRANCIA" if (case.legal_status or "").upper() == "APREHENDIDO" else "INVESTIGACION"
    payload["existing_gallery_files"] = [f.original_name for f in (case.files or []) if getattr(f, "file_type", "") == "gallery_image"]
    payload["existing_video_files"] = [f.original_name for f in (case.files or []) if getattr(f, "file_type", "") == "case_video"]
    payload["existing_annex_files"] = [f.original_name for f in (case.files or []) if getattr(f, "file_type", "") in ("novelty_attachment", "merchandise_photo", "denuncia_pdf", "diligencias_pdf", "audiencia_pdf")]
    payload["proceedings_payload"] = proceedings_from_storage(case.proceedings)
    return payload



def load_store_catalog():
    excel_path = BASE_DIR / "lista de tias_geo operaciones.xlsx"
    rows = []
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None or row[2] is None:
                continue
            vals = list(row) + [None] * max(0, 18 - len(row))
            rows.append({
                "central_code": "" if vals[0] is None else str(vals[0]).strip(),
                "central_name": "" if vals[1] is None else str(vals[1]).strip(),
                "store_code": "" if vals[2] is None else str(vals[2]).strip(),
                "store_name": "" if vals[3] is None else str(vals[3]).strip(),
                "format": "" if vals[4] is None else str(vals[4]).strip(),
                "province": "" if vals[11] is None else str(vals[11]).strip(),
                "region": "" if vals[6] is None else str(vals[6]).strip(),
                "district": "" if vals[7] is None else str(vals[7]).strip(),
                "zone": "" if vals[8] is None else str(vals[8]).strip(),
                "lat": vals[9],
                "lng": vals[10],
                "canton": "" if vals[12] is None else str(vals[12]).strip(),
                "subzone": "" if vals[13] is None else str(vals[13]).strip(),
                "police_district": "" if vals[14] is None else str(vals[14]).strip(),
                "circuit": "" if vals[15] is None else str(vals[15]).strip(),
                "subcircuit": "" if vals[16] is None else str(vals[16]).strip(),
                "subcircuit_code": "" if vals[17] is None else str(vals[17]).strip(),
            })
    return rows

STORE_CATALOG = load_store_catalog()

def central_name_for(db: Session, central_code: str | None):
    if not central_code:
        return ""
    for row in STORE_CATALOG:
        if str(row.get("central_code")) == str(central_code):
            return row.get("central_name", "")
    base = db.query(Store).filter(Store.store_code == central_code).first()
    return base.store_name if base else ""

def central_options():
    seen = {}
    for row in STORE_CATALOG:
        code = row.get("central_code", "")
        if code and code not in seen:
            seen[code] = row.get("central_name", "")
    return [{"code": k, "name": v} for k, v in seen.items()]


def first_report_image(case: Case, db: Session):
    """Obtiene la imagen principal para cartilla/preinforme.
    Prioridad absoluta: Fotografía integral de la novedad.
    Si no existe, usa como respaldo la última imagen de galería.
    """
    novelty = db.query(CaseFile).filter(
        CaseFile.case_id == case.id,
        CaseFile.file_type == "novelty_attachment",
    ).order_by(CaseFile.id.desc()).first()
    if novelty and novelty.stored_name:
        path = UPLOAD_DIR / novelty.stored_name
        if path.exists():
            return str(path)


def preinforme_image_options(case: Case):
    options = []
    seen = set()
    for rec in sorted(list(case.files or []), key=lambda f: (f.id or 0)):
        if getattr(rec, "file_type", "") not in ("gallery_image", "novelty_attachment", "merchandise_photo"):
            continue
        stored = getattr(rec, "stored_name", "") or ""
        original = getattr(rec, "original_name", "") or stored
        if not stored or stored in seen:
            continue
        path = UPLOAD_DIR / stored
        if not path.exists():
            continue
        seen.add(stored)
        kind = {
            "gallery_image": "Seguimiento",
            "novelty_attachment": "Integral novedad",
            "merchandise_photo": "Mercadería",
        }.get(rec.file_type, "Imagen")
        options.append({
            "value": f"upload:{stored}",
            "label": f"{kind}: {original}",
            "file_type": rec.file_type,
            "preview_url": f"/uploads/{stored}",
        })

    try:
        merch_items = json.loads(case.merchandise_json or "[]")
        if not isinstance(merch_items, list):
            merch_items = []
    except Exception:
        merch_items = []

    for item_idx, item in enumerate(merch_items, start=1):
        if not isinstance(item, dict):
            continue
        fotos = item.get("fotos", []) or []
        nombre = str(item.get("nombre_producto") or item.get("descripcion") or item.get("producto") or f"Producto {item_idx}")
        for foto_idx, foto in enumerate(fotos, start=1):
            if not isinstance(foto, str) or not foto.strip():
                continue
            token = f"merch:{item_idx-1}:{foto_idx-1}"
            options.append({
                "value": token,
                "label": f"Mercadería: {nombre} · Foto {foto_idx}",
                "file_type": "merchandise_json_photo",
                "preview_url": foto,
            })
    return options

    gallery = db.query(CaseFile).filter(
        CaseFile.case_id == case.id,
        CaseFile.file_type == "gallery_image",
    ).order_by(CaseFile.id.desc()).first()
    if gallery and gallery.stored_name:
        path = UPLOAD_DIR / gallery.stored_name
        if path.exists():
            return str(path)

    return None

def summary_payload(db: Session):
    cases = db.query(Case).all()

    def ordered(counter_obj, preferred=None):
        preferred = preferred or []
        data = dict(counter_obj)
        ordered_dict = {}
        for key in preferred:
            if key in data:
                ordered_dict[key] = data.pop(key)
        for key, value in sorted(data.items(), key=lambda kv: (-kv[1], str(kv[0]))):
            ordered_dict[key] = value
        return ordered_dict

    location_counts = Counter([c.novelty_location or "SIN DATO" for c in cases])
    novelty_counts = Counter([c.novelty_type or c.process or "SIN DATO" for c in cases])
    affectation_counts = Counter([c.affectation or "SIN DATO" for c in cases])
    alert_counts = Counter([c.alert_state or "SIN DATO" for c in cases])
    store_format_counts = Counter([c.format or "SIN DATO" for c in cases])
    region_counts = Counter([c.region or "SIN DATO" for c in cases])
    subtype_counts = Counter([c.novelty_subtype or "SIN DATO" for c in cases])
    store_ranking = Counter([f"{c.store_code or 'S/COD'}_{c.store_name or 'SIN NOMBRE'}" for c in cases])

    affectation_counts = ordered(affectation_counts, ["Directa", "Indirecta"])
    novelty_counts = ordered(novelty_counts, ["DELITO", "CONTRAVENCIÓN", "RIESGO", "TRANSITO"])
    store_format_counts = ordered(store_format_counts, ["TIA", "TIA EXPRES", "TIA GO"])
    region_counts = ordered(region_counts)
    subtype_counts = ordered(subtype_counts)
    store_ranking = ordered(store_ranking)

    return {
        "total_cases": len(cases),
        "location_counts": dict(location_counts),
        "novelty_counts": dict(novelty_counts),
        "novelty_total": sum(novelty_counts.values()),
        "affectation_counts": dict(affectation_counts),
        "affectation_total": sum(affectation_counts.values()),
        "alert_counts": dict(alert_counts),
        "alert_total": sum(alert_counts.values()),
        "store_format_counts": dict(store_format_counts),
        "store_format_total": sum(store_format_counts.values()),
        "region_counts": dict(region_counts),
        "subtype_counts": dict(subtype_counts),
        "store_ranking": dict(store_ranking),
    }

def build_case_excel(case: Case, db: Session, output_path: str):
    template = BASE_DIR / "FICHA NOVEDDAES TIA.xlsx"
    if template.exists():
        wb = openpyxl.load_workbook(template)
        ws = wb[wb.sheetnames[0]]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja3"
    suspects = json.loads(case.suspects_json or "[]")
    first_suspect = suspects[0] if suspects else {}
    fills = {
        "B3": case.event_date or "",
        "B4": case.event_time or "",
        "B5": f"{case.store_code or ''}-{case.store_name or ''}",
        "B6": (case.novelty_location or "").upper(),
        "B7": (case.novelty_type or case.process or "").upper(),
        "B8": (case.novelty_subtype or "").upper(),
        "B9": (case.modality or "").upper(),
        "B10": case.number_of_members or "",
        "B11": first_suspect.get("movilizacion", "") or "",
        "B12": (first_suspect.get("arma_identificada", "") or "").upper(),
        "B13": case.subzone or "",
        "B14": case.police_district or "",
        "B15": case.circuit or "",
        "B16": case.subcircuit or "",
        "B17": f"({case.lat or ''}-{case.lng or ''})",
        "B18": first_suspect.get("actividad", "") or "",
        "B19": case.detail or "",
    }
    for cell, value in fills.items():
        ws[cell] = value
    # summary sheet
    if "Resumen" in wb.sheetnames:
        del wb["Resumen"]
    rs = wb.create_sheet("Resumen")
    summary = summary_payload(db)
    rs["A1"] = "REPORTE EJECUTIVO GDN"
    rs["A3"] = "TOTAL CASOS"; rs["A4"]="INTERIOR"; rs["B4"]=summary["location_counts"].get("Interior del local",0); rs["A5"]="EXTERIOR"; rs["B5"]=summary["location_counts"].get("Exterior del local",0); rs["A6"]="TOTAL"; rs["B6"]=summary["total_cases"]
    rs["D3"] = "NOVEDAD"
    r = 4
    for k,v in summary["novelty_counts"].items():
        rs[f"D{r}"]=k; rs[f"E{r}"]=v; r+=1
    rs[f"D{r}"]="TOTAL"; rs[f"E{r}"]=sum(summary["novelty_counts"].values())
    rs["A8"]="AFECTACIÓN"
    r=9
    for k,v in summary["affectation_counts"].items():
        rs[f"A{r}"]=k; rs[f"B{r}"]=v; r+=1
    rs[f"A{r}"]="TOTAL"; rs[f"B{r}"]=summary["affectation_total"]
    rs["A13"]="ALERTAS"
    r=14
    for k,v in summary["alert_counts"].items():
        rs[f"A{r}"]=k; rs[f"B{r}"]=v; r+=1
    rs[f"A{r}"]="TOTAL"; rs[f"B{r}"]=summary["alert_total"]
    rs["A18"]="TIENDAS"
    r=19
    for k,v in summary["store_format_counts"].items():
        rs[f"A{r}"]=k; rs[f"B{r}"]=v; r+=1
    rs[f"A{r}"]="TOTAL"; rs[f"B{r}"]=summary["store_format_total"]
    rs["A24"]="REGIÓN"
    c=1
    for name, count in summary["region_counts"].items():
        rs.cell(25, c, name)
        rs.cell(26, c, count)
        c += 1
    wb.save(output_path)

def common_context(user):
    return {
        "user": user,
        "process_options": PROCESS_OPTIONS,
        "subprocess_options": SUBPROCESS_OPTIONS,
        "security_companies": SECURITY_COMPANIES,
        "alert_states": ALERT_STATES,
        "discarded_reasons": DISCARDED_REASONS,
        "affectations": AFFECTATIONS,
        "novelty_locations": NOVELTY_LOCATIONS,
        "places": PLACES,
        "place_subtypes": PLACE_SUBTYPES,
        "police_units": POLICE_UNITS,
        "requested_by_options": REQUESTED_BY,
        "novelty_types": NOVELTY_TYPES,
        "suspect_status_options": globals().get("SUSPECT_STATUS", ["Aprehendido", "No aprehendido", "Procesado"]),
        "antecedentes_options": globals().get("ANTECEDENTES", ["Reincidente", "Sospechoso no reconocido"]),
        "tipos_organizacion": globals().get("TIPOS_ORGANIZACION", ["Delincuencia común", "Delincuencia organizada", "Delincuente habitual"]),
        "org_names": globals().get("ORG_NAMES", ["No reconocida"]),
        "nationalities": globals().get("NATIONALITIES", ["Ecuador"]),
        "product_types": PRODUCT_TYPES,
        "merch_states": MERCH_STATES,
        "legal_statuses": LEGAL_STATUSES,
        "precautionary_measures": PRECAUTIONARY_MEASURES,
        "novelty_matrix": NOVELTY_MATRIX,
        "modalidad_catalogo": MODALIDAD_CATALOGO,
    }


@app.get("/", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if user:
        return RedirectResponse("/dashboard", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "error": None, "user": None})


@app.post("/login", response_class=HTMLResponse)
def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == username, User.password == password).first()
    if not user:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Credenciales inválidas", "user": None})
    request.session["user_id"] = user.id
    return RedirectResponse("/dashboard", status_code=302)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/", status_code=302)


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    case_count = db.query(func.count(Case.id)).scalar() or 0
    confirmed_count = db.query(func.count(Case.id)).filter(Case.alert_state == "Confirmado").scalar() or 0
    stores_count = db.query(func.count(Store.id)).scalar() or 0
    recent_cases = db.query(Case).order_by(Case.created_at.desc()).limit(10).all()
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": user,
        "case_count": case_count,
        "confirmed_count": confirmed_count,
        "stores_count": stores_count,
        "recent_cases": recent_cases,
    })



@app.get("/cases/new", response_class=HTMLResponse)
def new_case(request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    stores = STORE_CATALOG if STORE_CATALOG else [{
        "store_code": s.store_code,
        "central_code": s.central_code,
        "central_name": central_name_for(db, s.central_code),
        "store_name": s.store_name,
        "format": s.format,
        "region": s.region,
        "district": s.district,
        "zone": s.zone,
        "lat": s.lat,
        "lng": s.lng,
        "province": s.province,
        "canton": s.canton,
        "subzone": s.subzone,
        "police_district": s.police_district,
        "circuit": s.circuit,
        "subcircuit": s.subcircuit,
        "subcircuit_code": s.subcircuit_code,
    } for s in db.query(Store).order_by(Store.store_code).all()]
    ctx = common_context(user)
    ctx.update({
        "request": request,
        "stores": [SimpleNamespace(**row) for row in stores],
        "stores_json": stores,
        "central_options": central_options(),
        "case_data": None,
        "form_action": "/cases/new",
        "submit_label": "Guardar caso",
        "page_mode": "create",
    })
    return templates.TemplateResponse("case_form.html", ctx)



@app.post("/cases/new")
async def create_case(
    request: Request,
    police_unit: str = Form(""),
    requested_by: str = Form(""),
    process: str = Form(""),
    sub_process: str = Form(""),
    alarm_activated: str = Form("NO"),
    chief_name: str = Form(""),
    alert_time: str = Form(""),
    security_company: str = Form(""),
    security_operator: str = Form(""),
    center_operator_time: str = Form(""),
    center_operator_name: str = Form(""),
    central_code: str = Form(""),
    store_code: str = Form(...),
    affectation: str = Form(""),
    novelty_location: str = Form(""),
    place_company: str = Form(""),
    place_subtype: str = Form(""),
    event_date: str = Form(""),
    event_time: str = Form(""),
    alert_state: str = Form(""),
    discarded_reason: str = Form(""),
    novelty_type: str = Form(""),
    novelty_subtype: str = Form(""),
    modality: str = Form(""),
    number_of_members: int = Form(1),
    detail: str = Form(""),
    victims_json: str = Form("[]"),
    suspects_json: str = Form("[]"),
    merchandise_json: str = Form("[]"),
    photo_notes_json: str = Form("[]"),
    vehicles_json: str = Form("[]"),
    processed_json: str = Form("[]"),
    has_merchandise: str = Form("NO"),
    proceedings_json: str = Form("[]"),
    delegation: str = Form(""),
    prosecutor_office: str = Form(""),
    prosecutor_name: str = Form(""),
    complaint_number: str = Form(""),
    complaint_date: str = Form(""),
    judicial_unit: str = Form(""),
    cause_number: str = Form(""),
    proceedings: str = Form(""),
    legal_status: str = Form(""),
    hearing_date: str = Form(""),
    precautionary_measure: str = Form(""),
    legal_summary: str = Form(""),
    gallery_1: UploadFile | None = File(None),
    gallery_2: UploadFile | None = File(None),
    gallery_3: UploadFile | None = File(None),
    gallery_4: UploadFile | None = File(None),
    gallery_5: UploadFile | None = File(None),
    gallery_6: UploadFile | None = File(None),
    gallery_7: UploadFile | None = File(None),
    gallery_8: UploadFile | None = File(None),
    gallery_9: UploadFile | None = File(None),
    gallery_10: UploadFile | None = File(None),
    gallery_11: UploadFile | None = File(None),
    gallery_12: UploadFile | None = File(None),
    gallery_13: UploadFile | None = File(None),
    gallery_14: UploadFile | None = File(None),
    gallery_15: UploadFile | None = File(None),
    gallery_16: UploadFile | None = File(None),
    gallery_17: UploadFile | None = File(None),
    gallery_18: UploadFile | None = File(None),
    gallery_19: UploadFile | None = File(None),
    gallery_20: UploadFile | None = File(None),
    merchandise_photo: UploadFile | None = File(None),
    novelty_attachment: UploadFile | None = File(None),
    denuncia_pdf: UploadFile | None = File(None),
    diligencias_pdf: UploadFile | None = File(None),
    audiencia_pdf: UploadFile | None = File(None),
    case_video: List[UploadFile] = File(default=[]),
    db: Session = Depends(get_db)
):
    user = require_login(request, db)
    store = db.query(Store).filter(Store.store_code == store_code).first()
    if not store:
        raise HTTPException(status_code=404, detail="Tienda no encontrada")

    process_value = process if alarm_activated == "SI" else ""
    base_code = central_code if process_value == "HURTO" and central_code else store_code
    case_number = build_case_number(db, base_code)
    case = Case(
        case_number=case_number,
        police_unit=police_unit,
        requested_by=requested_by,
        process=process_value,
        sub_process=sub_process if alarm_activated == "SI" else "",
        alarm_activated=alarm_activated,
        chief_name=chief_name if alarm_activated == "SI" else "",
        alert_time=alert_time if alarm_activated == "SI" else "",
        security_company=security_company if alarm_activated == "SI" else "",
        security_operator=security_operator if alarm_activated == "SI" else "",
        center_operator_time=center_operator_time if alarm_activated == "SI" else "",
        center_operator_name=center_operator_name if alarm_activated == "SI" else "",
        central_code=central_code if process_value == "HURTO" else "",
        central_name=central_name_for(db, central_code) if process_value == "HURTO" else "",
        store_code=store.store_code,
        store_name=store.store_name,
        format=store.format,
        region=store.region,
        district=store.district,
        zone=store.zone,
        province=store.province,
        canton=store.canton,
        parish=store.parish,
        subzone=store.subzone,
        police_district=store.police_district,
        circuit=store.circuit,
        subcircuit=store.subcircuit,
        subcircuit_code=store.subcircuit_code,
        lat=str(store.lat or ""),
        lng=str(store.lng or ""),
        affectation=affectation,
        novelty_location=novelty_location,
        place_company=place_company,
        place_subtype=place_subtype,
        event_date=event_date,
        event_time=event_time,
        alert_state=alert_state,
        discarded_reason=discarded_reason,
        novelty_type=novelty_type,
        novelty_subtype=novelty_subtype,
        modality=modality,
        number_of_members=number_of_members,
        detail=detail,
        victims_json=to_json(victims_json),
        suspects_json=to_json(suspects_json),
        merchandise_json=to_json(merchandise_json) if has_merchandise == "SI" else "[]",
        photo_notes_json=to_json(photo_notes_json),
        vehicles_json=to_json(vehicles_json),
        processed_json=to_json(processed_json),
        delegation=delegation,
        prosecutor_office=prosecutor_office,
        prosecutor_name=prosecutor_name,
        complaint_number=complaint_number,
        complaint_date=complaint_date,
        judicial_unit=judicial_unit,
        cause_number=cause_number,
        proceedings=proceedings_to_storage(proceedings, proceedings_json),
        legal_status=legal_status,
        hearing_date=hearing_date,
        precautionary_measure=precautionary_measure,
        legal_summary=legal_summary,
        created_by=user.id,
    )
    db.add(case)
    db.commit()
    db.refresh(case)

    uploads = [
        gallery_1, gallery_2, gallery_3, gallery_4, gallery_5,
        gallery_6, gallery_7, gallery_8, gallery_9, gallery_10,
        gallery_11, gallery_12, gallery_13, gallery_14, gallery_15,
        gallery_16, gallery_17, gallery_18, gallery_19, gallery_20,
    ]
    for up in uploads:
        save_upload(up, "gallery_image", case.id, db)
    for up, file_type in [
        (merchandise_photo, "merchandise_photo"),
        (novelty_attachment, "novelty_attachment"),
        (denuncia_pdf, "denuncia_pdf"),
        (diligencias_pdf, "diligencias_pdf"),
        (audiencia_pdf, "audiencia_pdf"),
    ]:
        save_upload(up, file_type, case.id, db)
    for up in (case_video or []):
        save_upload(up, "case_video", case.id, db)
    db.commit()
    return RedirectResponse(f"/cases/{case.id}", status_code=302)


@app.get("/cases", response_class=HTMLResponse)
def list_cases(request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    cases = db.query(Case).order_by(Case.created_at.desc()).all()
    return templates.TemplateResponse("case_list.html", {"request": request, "user": user, "cases": cases})


@app.get("/cases/{case_id}", response_class=HTMLResponse)
def case_detail(case_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404)
    def safe_json_list(raw):
        try:
            data = json.loads(raw or "[]")
            return data if isinstance(data, list) else []
        except Exception:
            return []

    return templates.TemplateResponse("case_detail.html", {
        "request": request,
        "user": user,
        "case": case,
        "victims": safe_json_list(case.victims_json),
        "suspects": safe_json_list(case.suspects_json),
        "merch": safe_json_list(case.merchandise_json),
        "photo_notes": safe_json_list(case.photo_notes_json),
        "vehicles": safe_json_list(case.vehicles_json),
        "processed": safe_json_list(case.processed_json),
        "proceedings_payload": proceedings_from_storage(case.proceedings),
        "case_files": case.files,
    })


@app.get("/cases/{case_id}/pdf")
def case_pdf(case_id: int, request: Request, db: Session = Depends(get_db)):
    require_login(request, db)
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404)
    try:
        pdf_path = generate_case_pdf(case, str(PDF_DIR))
    except Exception:
        payload = case_to_form_data(case)
        case.victims_json = json.dumps(payload.get("victims", []), ensure_ascii=False)
        case.suspects_json = json.dumps(payload.get("suspects", []), ensure_ascii=False)
        case.merchandise_json = json.dumps(payload.get("merchandise", []), ensure_ascii=False)
        case.vehicles_json = json.dumps(payload.get("vehicles", []), ensure_ascii=False)
        case.processed_json = json.dumps(payload.get("processed", []), ensure_ascii=False)
        case.photo_notes_json = json.dumps(payload.get("photo_notes", []), ensure_ascii=False)
        pdf_path = generate_case_pdf(case, str(PDF_DIR))
    return FileResponse(pdf_path, media_type="application/pdf", filename=f"{case.case_number}.pdf")




@app.get("/cases/{case_id}/edit", response_class=HTMLResponse)
def edit_case(case_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_admin(request, db)
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404)
    stores = STORE_CATALOG if STORE_CATALOG else [{
        "store_code": s.store_code,
        "central_code": s.central_code,
        "central_name": central_name_for(db, s.central_code),
        "store_name": s.store_name,
        "format": s.format,
        "region": s.region,
        "district": s.district,
        "zone": s.zone,
        "lat": s.lat,
        "lng": s.lng,
        "province": s.province,
        "canton": s.canton,
        "subzone": s.subzone,
        "police_district": s.police_district,
        "circuit": s.circuit,
        "subcircuit": s.subcircuit,
        "subcircuit_code": s.subcircuit_code,
    } for s in db.query(Store).order_by(Store.store_code).all()]
    ctx = common_context(user)
    ctx.update({
        "request": request,
        "stores": [SimpleNamespace(**row) for row in stores],
        "stores_json": stores,
        "central_options": central_options(),
        "case_data": case_to_form_data(case),
        "form_action": f"/cases/{case.id}/edit",
        "submit_label": "Guardar cambios",
        "page_mode": "edit",
    })
    return templates.TemplateResponse("case_form.html", ctx)

@app.post("/cases/{case_id}/edit")
async def update_case(
    case_id: int,
    request: Request,
    police_unit: str = Form(""),
    requested_by: str = Form(""),
    process: str = Form(""),
    sub_process: str = Form(""),
    alarm_activated: str = Form("NO"),
    chief_name: str = Form(""),
    alert_time: str = Form(""),
    security_company: str = Form(""),
    security_operator: str = Form(""),
    center_operator_time: str = Form(""),
    center_operator_name: str = Form(""),
    central_code: str = Form(""),
    store_code: str = Form(...),
    affectation: str = Form(""),
    novelty_location: str = Form(""),
    place_company: str = Form(""),
    place_subtype: str = Form(""),
    event_date: str = Form(""),
    event_time: str = Form(""),
    alert_state: str = Form(""),
    discarded_reason: str = Form(""),
    novelty_type: str = Form(""),
    novelty_subtype: str = Form(""),
    modality: str = Form(""),
    number_of_members: int = Form(1),
    detail: str = Form(""),
    victims_json: str = Form("[]"),
    suspects_json: str = Form("[]"),
    merchandise_json: str = Form("[]"),
    photo_notes_json: str = Form("[]"),
    vehicles_json: str = Form("[]"),
    processed_json: str = Form("[]"),
    has_merchandise: str = Form("NO"),
    proceedings_json: str = Form("[]"),
    delegation: str = Form(""),
    prosecutor_office: str = Form(""),
    prosecutor_name: str = Form(""),
    complaint_number: str = Form(""),
    complaint_date: str = Form(""),
    judicial_unit: str = Form(""),
    cause_number: str = Form(""),
    proceedings: str = Form(""),
    legal_status: str = Form(""),
    hearing_date: str = Form(""),
    precautionary_measure: str = Form(""),
    legal_summary: str = Form(""),
    gallery_1: UploadFile | None = File(None),
    gallery_2: UploadFile | None = File(None),
    gallery_3: UploadFile | None = File(None),
    gallery_4: UploadFile | None = File(None),
    gallery_5: UploadFile | None = File(None),
    gallery_6: UploadFile | None = File(None),
    gallery_7: UploadFile | None = File(None),
    gallery_8: UploadFile | None = File(None),
    gallery_9: UploadFile | None = File(None),
    gallery_10: UploadFile | None = File(None),
    gallery_11: UploadFile | None = File(None),
    gallery_12: UploadFile | None = File(None),
    gallery_13: UploadFile | None = File(None),
    gallery_14: UploadFile | None = File(None),
    gallery_15: UploadFile | None = File(None),
    gallery_16: UploadFile | None = File(None),
    gallery_17: UploadFile | None = File(None),
    gallery_18: UploadFile | None = File(None),
    gallery_19: UploadFile | None = File(None),
    gallery_20: UploadFile | None = File(None),
    merchandise_photo: UploadFile | None = File(None),
    novelty_attachment: UploadFile | None = File(None),
    denuncia_pdf: UploadFile | None = File(None),
    diligencias_pdf: UploadFile | None = File(None),
    audiencia_pdf: UploadFile | None = File(None),
    case_video: List[UploadFile] = File(default=[]),
    db: Session = Depends(get_db)
):
    require_admin(request, db)
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404)
    store = db.query(Store).filter(Store.store_code == store_code).first()
    if not store:
        raise HTTPException(status_code=404, detail="Tienda no encontrada")

    save_case_backup(case)
    process_value = process if alarm_activated == "SI" else case.process
    case.police_unit = police_unit
    case.requested_by = requested_by
    case.process = process
    case.sub_process = sub_process
    case.alarm_activated = alarm_activated
    case.chief_name = chief_name if alarm_activated == "SI" else ""
    case.alert_time = alert_time if alarm_activated == "SI" else ""
    case.security_company = security_company if alarm_activated == "SI" else ""
    case.security_operator = security_operator if alarm_activated == "SI" else ""
    case.center_operator_time = center_operator_time if alarm_activated == "SI" else ""
    case.center_operator_name = center_operator_name if alarm_activated == "SI" else ""
    case.central_code = central_code if process_value == "HURTO" else ""
    case.central_name = central_name_for(db, central_code) if process_value == "HURTO" else ""
    case.store_code = store.store_code
    case.store_name = store.store_name
    case.format = store.format
    case.region = store.region
    case.district = store.district
    case.zone = store.zone
    case.province = store.province
    case.canton = store.canton
    case.parish = store.parish
    case.subzone = store.subzone
    case.police_district = store.police_district
    case.circuit = store.circuit
    case.subcircuit = store.subcircuit
    case.subcircuit_code = store.subcircuit_code
    case.lat = str(store.lat or "")
    case.lng = str(store.lng or "")
    case.affectation = affectation
    case.novelty_location = novelty_location
    case.place_company = place_company
    case.place_subtype = place_subtype
    case.event_date = event_date
    case.event_time = event_time
    case.alert_state = alert_state
    case.discarded_reason = discarded_reason
    case.novelty_type = novelty_type
    case.novelty_subtype = novelty_subtype
    case.modality = modality
    case.number_of_members = number_of_members
    case.detail = detail
    case.victims_json = to_json(victims_json)
    case.suspects_json = to_json(suspects_json)
    case.merchandise_json = to_json(merchandise_json) if has_merchandise == "SI" else "[]"
    case.photo_notes_json = to_json(photo_notes_json)
    case.vehicles_json = to_json(vehicles_json)
    case.processed_json = to_json(processed_json)
    case.delegation = delegation
    case.prosecutor_office = prosecutor_office
    case.prosecutor_name = prosecutor_name
    case.complaint_number = complaint_number
    case.complaint_date = complaint_date
    case.judicial_unit = judicial_unit
    case.cause_number = cause_number
    case.proceedings = proceedings_to_storage(proceedings, proceedings_json)
    case.legal_status = legal_status
    case.hearing_date = hearing_date
    case.precautionary_measure = precautionary_measure
    case.legal_summary = legal_summary

    uploads = [
        gallery_1, gallery_2, gallery_3, gallery_4, gallery_5,
        gallery_6, gallery_7, gallery_8, gallery_9, gallery_10,
        gallery_11, gallery_12, gallery_13, gallery_14, gallery_15,
        gallery_16, gallery_17, gallery_18, gallery_19, gallery_20,
    ]
    for up in uploads:
        save_upload(up, "gallery_image", case.id, db)
    for up, file_type in [
        (merchandise_photo, "merchandise_photo"),
        (novelty_attachment, "novelty_attachment"),
        (denuncia_pdf, "denuncia_pdf"),
        (diligencias_pdf, "diligencias_pdf"),
        (audiencia_pdf, "audiencia_pdf"),
    ]:
        save_upload(up, file_type, case.id, db)
    for up in (case_video or []):
        save_upload(up, "case_video", case.id, db)

    db.commit()
    return RedirectResponse(f"/cases/{case.id}", status_code=302)

@app.get("/buscador-ia", response_class=HTMLResponse)
def buscador_ia(request: Request, q_persona: str = "", q_vehiculo: str = "", db: Session = Depends(get_db)):
    user = require_login(request, db)
    q_persona_norm = (q_persona or "").strip().lower()
    q_vehiculo_norm = (q_vehiculo or "").strip().lower()

    def safe_list(raw):
        try:
            data = json.loads(raw or "[]")
            return data if isinstance(data, list) else []
        except Exception:
            return []

    person_summary = {"cedula":"", "nombre":"", "organizacion":"", "caso":""}
    vehicle_summary = {"tipo":"", "marca":"", "placa":"", "color":""}
    person_cases = []
    linked_people = []
    vehicle_cases = []

    cases = db.query(Case).order_by(Case.created_at.desc()).all()
    for case in cases:
        suspects = safe_list(case.suspects_json)
        vehicles = safe_list(case.vehicles_json)

        matched_persons = []
        matched_vehicles = []

        if q_persona_norm:
            for s in suspects:
                haystack = " ".join([
                    str(s.get("cedula", "")),
                    str(s.get("nombre", "")),
                    str(s.get("organizacion_nombre", "")),
                    str(s.get("antecedentes", "")),
                    str(s.get("situacion", "")),
                    str(case.novelty_subtype or ""),
                    str(case.store_code or ""),
                    str(case.store_name or ""),
                    str(case.subzone or ""),
                ]).lower()
                if q_persona_norm in haystack:
                    matched_persons.append(s)

        if q_vehiculo_norm:
            for v in vehicles:
                haystack = " ".join([
                    str(v.get("tipo_vehiculo", "")),
                    str(v.get("marca", "")),
                    str(v.get("placas", "")),
                    str(v.get("color", "")),
                    str(v.get("caracteristicas", "")),
                    str(case.novelty_subtype or ""),
                ]).lower()
                if q_vehiculo_norm in haystack:
                    matched_vehicles.append(v)

        if matched_persons:
            if not person_summary["nombre"]:
                s = matched_persons[0]
                person_summary = {
                    "cedula": s.get("cedula", ""),
                    "nombre": s.get("nombre", ""),
                    "organizacion": s.get("organizacion_nombre", ""),
                    "caso": case.case_number,
                }
            for s in matched_persons:
                person_cases.append({
                    "case_number": case.case_number,
                    "nombre": s.get("nombre", ""),
                    "cedula": s.get("cedula", ""),
                    "sexo": s.get("sexo", ""),
                    "edad": s.get("edad", ""),
                    "antecedentes": s.get("antecedentes", ""),
                    "situacion": s.get("situacion", ""),
                    "organizacion": s.get("organizacion_nombre", ""),
                    "subtipo": case.novelty_subtype or "",
                    "tienda": f"{case.store_code or ''}-{case.store_name or ''}",
                    "subzona": case.subzone or "",
                })
            for other_case in cases:
                for s2 in safe_list(other_case.suspects_json):
                    if person_summary["nombre"] and str(s2.get("nombre", "")).strip().lower() == str(person_summary["nombre"]).strip().lower() and other_case.case_number != case.case_number:
                        linked_people.append({
                            "case_number": other_case.case_number,
                            "nombre": s2.get("nombre", ""),
                            "cedula": s2.get("cedula", ""),
                            "sexo": s2.get("sexo", ""),
                            "edad": s2.get("edad", ""),
                            "antecedentes": s2.get("antecedentes", ""),
                            "situacion": s2.get("situacion", ""),
                            "organizacion": s2.get("organizacion_nombre", ""),
                            "subtipo": other_case.novelty_subtype or "",
                            "tienda": f"{other_case.store_code or ''}-{other_case.store_name or ''}",
                            "subzona": other_case.subzone or "",
                        })

        if matched_vehicles:
            if not vehicle_summary["placa"]:
                v = matched_vehicles[0]
                vehicle_summary = {
                    "tipo": v.get("tipo_vehiculo", ""),
                    "marca": v.get("marca", ""),
                    "placa": v.get("placas", ""),
                    "color": v.get("color", ""),
                }
            for v in matched_vehicles:
                vehicle_cases.append({
                    "case_number": case.case_number,
                    "tipo": v.get("tipo_vehiculo", ""),
                    "marca": v.get("marca", ""),
                    "placas": v.get("placas", ""),
                    "color": v.get("color", ""),
                    "subtipo": case.novelty_subtype or "",
                    "actividad": v.get("actividades_sospechosas", ""),
                    "caracteristicas": v.get("caracteristicas", ""),
                })

    def dedupe(rows, key_fields):
        seen = set()
        out = []
        for row in rows:
            k = tuple(str(row.get(f, "")) for f in key_fields)
            if k in seen:
                continue
            seen.add(k)
            out.append(row)
        return out

    person_cases = dedupe(person_cases, ["case_number", "nombre", "cedula"])
    linked_people = dedupe(linked_people, ["case_number", "nombre", "cedula"])
    vehicle_cases = dedupe(vehicle_cases, ["case_number", "placas", "marca"])

    return templates.TemplateResponse("buscador_ia.html", {
        "request": request,
        "user": user,
        "q_persona": q_persona,
        "q_vehiculo": q_vehiculo,
        "person_summary": person_summary,
        "vehicle_summary": vehicle_summary,
        "person_cases": person_cases,
        "linked_people": linked_people,
        "vehicle_cases": vehicle_cases,
    })

@app.get("/reports", response_class=HTMLResponse)
def reports_page(request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    summary = summary_payload(db)
    recent_cases = db.query(Case).order_by(Case.created_at.desc()).limit(15).all()
    return templates.TemplateResponse("reports.html", {
        "request": request,
        "user": user,
        "summary": summary,
        "recent_cases": recent_cases,
    })

@app.get("/reports/summary/pdf")
def reports_summary_pdf(request: Request, db: Session = Depends(get_db)):
    require_login(request, db)
    summary = summary_payload(db)
    pdf_path = generate_summary_pdf(summary, str(PDF_DIR))
    return FileResponse(pdf_path, media_type="application/pdf", filename="reporte_resumen_tracker.pdf")

@app.get("/reports/summary/excel")
def reports_summary_excel(request: Request, db: Session = Depends(get_db)):
    require_login(request, db)
    latest_case = db.query(Case).order_by(Case.created_at.desc()).first()
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    if latest_case:
        build_case_excel(latest_case, db, tmp.name)
    else:
        wb = openpyxl.Workbook()
        wb.save(tmp.name)
    return FileResponse(tmp.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="reporte_tracker.xlsx")


@app.get("/reports/executive/pdf")
def reports_executive_pdf(request: Request, db: Session = Depends(get_db)):
    require_login(request, db)
    summary = summary_payload(db)
    pdf_path = generate_executive_dashboard_pdf(summary, str(PDF_DIR))
    return FileResponse(pdf_path, media_type="application/pdf", filename="reporte_ejecutivo_gdn.pdf")

@app.get("/reports/{case_id}/pdf")
def reports_case_pdf(case_id: int, request: Request, db: Session = Depends(get_db)):
    require_login(request, db)
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404)
    image_path = first_report_image(case, db)
    try:
        pdf_path = generate_case_sheet_pdf(case, str(PDF_DIR), image_path=image_path)
    except Exception:
        payload = case_to_form_data(case)
        case.suspects_json = json.dumps(payload.get("suspects", []), ensure_ascii=False)
        pdf_path = generate_case_sheet_pdf(case, str(PDF_DIR), image_path=image_path)
    return FileResponse(pdf_path, media_type="application/pdf", filename=f"{case.case_number}_ficha.pdf")



@app.get("/reports/{case_id}/preinforme/preview", response_class=HTMLResponse)
def reports_case_preinforme_preview(case_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404)
    default_actions = ""
    image_options = preinforme_image_options(case)
    default_annexes = []
    if image_options:
        default_annexes.append({"title": "fotografía integral de la novedad", "source": image_options[0]["value"]})
    if len(image_options) > 1:
        default_annexes.append({"title": "fotografía de mercadería no recuperada", "source": image_options[1]["value"]})
    return templates.TemplateResponse("preinforme_preview.html", {
        "request": request,
        "user": user,
        "case": case,
        "default_asunto": case.novelty_subtype or case.novelty_type or "",
        "default_detalle_asunto": " / ".join([v for v in [case.novelty_type or "", case.novelty_subtype or "", case.modality or ""] if v]),
        "default_acciones": default_actions,
        "image_options": image_options,
        "default_annexes": default_annexes,
    })

@app.get("/reports/{case_id}/preinforme/view")
def reports_case_preinforme_view(case_id: int, request: Request, asunto: str = "", detalle_asunto: str = "", acciones_realizadas: str = "", db: Session = Depends(get_db)):
    user = require_login(request, db)
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404)

    def _existing_path(stored_name: str):
        p = UPLOAD_DIR / stored_name
        return str(p) if p.exists() else None

    image_path = first_report_image(case, db)
    gallery_images = []
    annex_images = []

    files = sorted(list(case.files or []), key=lambda f: (f.id or 0))
    for rec in files:
        p = _existing_path(rec.stored_name)
        if not p:
            continue
        if rec.file_type == "gallery_image":
            gallery_images.append(p)
        elif rec.file_type == "novelty_attachment":
            annex_images.insert(0, p)
        elif rec.file_type == "merchandise_photo":
            annex_images.append(p)

    analyst_name = ""
    if getattr(case, "creator", None) and getattr(case.creator, "full_name", None):
        analyst_name = case.creator.full_name
    elif getattr(user, "full_name", None):
        analyst_name = user.full_name

    titles = request.query_params.getlist("annex_title")
    sources = request.query_params.getlist("annex_source")
    selected_annexes = []
    try:
        merch_items = json.loads(case.merchandise_json or "[]")
        if not isinstance(merch_items, list):
            merch_items = []
    except Exception:
        merch_items = []
    for idx, src_name in enumerate(sources):
        src_name = (src_name or "").strip()
        if not src_name:
            continue
        resolved = None
        if src_name.startswith("upload:"):
            stored_name = src_name.split(":", 1)[1]
            path = UPLOAD_DIR / stored_name
            if path.exists():
                resolved = {"kind": "path", "value": str(path)}
        elif src_name.startswith("merch:"):
            try:
                _, item_i, foto_i = src_name.split(":")
                item = merch_items[int(item_i)]
                fotos = item.get("fotos", []) or []
                foto = fotos[int(foto_i)]
                if isinstance(foto, str) and foto.strip():
                    resolved = {"kind": "data_url", "value": foto}
            except Exception:
                resolved = None
        else:
            path = UPLOAD_DIR / src_name
            if path.exists():
                resolved = {"kind": "path", "value": str(path)}
        if not resolved:
            continue
        title = titles[idx].strip() if idx < len(titles) and titles[idx] else f"Anexo {idx+1}"
        selected_annexes.append({"title": title, "path": resolved["value"], "source_kind": resolved["kind"]})
    if not selected_annexes:
        for idx, path in enumerate(annex_images[:2], start=1):
            selected_annexes.append({"title": f"Anexo {idx}", "path": path})

    pdf_path = generate_preinforme_pdf(
        case,
        str(PDF_DIR),
        image_path=image_path,
        seguimiento_images=gallery_images,
        annex_images=annex_images,
        analyst_name=analyst_name,
        asunto=asunto,
        detalle_asunto=detalle_asunto,
        acciones_realizadas=acciones_realizadas,
        selected_annexes=selected_annexes,
    )
    return FileResponse(pdf_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=%s_preinforme.pdf" % (case.case_number,)})

@app.get("/reports/{case_id}/preinforme")
def reports_case_preinforme(case_id: int, request: Request, asunto: str = "", detalle_asunto: str = "", acciones_realizadas: str = "", db: Session = Depends(get_db)):
    user = require_login(request, db)
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404)

    def _existing_path(stored_name: str):
        p = UPLOAD_DIR / stored_name
        return str(p) if p.exists() else None

    image_path = first_report_image(case, db)
    gallery_images = []
    annex_images = []

    files = sorted(list(case.files or []), key=lambda f: (f.id or 0))
    for rec in files:
        p = _existing_path(rec.stored_name)
        if not p:
            continue
        if rec.file_type == "gallery_image":
            gallery_images.append(p)
        elif rec.file_type == "novelty_attachment":
            annex_images.insert(0, p)
        elif rec.file_type == "merchandise_photo":
            annex_images.append(p)

    analyst_name = ""
    if getattr(case, "creator", None) and getattr(case.creator, "full_name", None):
        analyst_name = case.creator.full_name
    elif getattr(user, "full_name", None):
        analyst_name = user.full_name

    titles = request.query_params.getlist("annex_title")
    sources = request.query_params.getlist("annex_source")
    selected_annexes = []
    try:
        merch_items = json.loads(case.merchandise_json or "[]")
        if not isinstance(merch_items, list):
            merch_items = []
    except Exception:
        merch_items = []
    for idx, src_name in enumerate(sources):
        src_name = (src_name or "").strip()
        if not src_name:
            continue
        resolved = None
        if src_name.startswith("upload:"):
            stored_name = src_name.split(":", 1)[1]
            path = UPLOAD_DIR / stored_name
            if path.exists():
                resolved = {"kind": "path", "value": str(path)}
        elif src_name.startswith("merch:"):
            try:
                _, item_i, foto_i = src_name.split(":")
                item = merch_items[int(item_i)]
                fotos = item.get("fotos", []) or []
                foto = fotos[int(foto_i)]
                if isinstance(foto, str) and foto.strip():
                    resolved = {"kind": "data_url", "value": foto}
            except Exception:
                resolved = None
        else:
            path = UPLOAD_DIR / src_name
            if path.exists():
                resolved = {"kind": "path", "value": str(path)}
        if not resolved:
            continue
        title = titles[idx].strip() if idx < len(titles) and titles[idx] else f"Anexo {idx+1}"
        selected_annexes.append({"title": title, "path": resolved["value"], "source_kind": resolved["kind"]})
    if not selected_annexes:
        for idx, path in enumerate(annex_images[:2], start=1):
            selected_annexes.append({"title": f"Anexo {idx}", "path": path})

    pdf_path = generate_preinforme_pdf(
        case,
        str(PDF_DIR),
        image_path=image_path,
        seguimiento_images=gallery_images,
        annex_images=annex_images,
        analyst_name=analyst_name,
        asunto=asunto,
        detalle_asunto=detalle_asunto,
        acciones_realizadas=acciones_realizadas,
        selected_annexes=selected_annexes,
    )
    return FileResponse(pdf_path, media_type="application/pdf", filename=f"{case.case_number}_preinforme.pdf", headers={"Content-Disposition": f"attachment; filename={case.case_number}_preinforme.pdf"})

@app.get("/reports/{case_id}/excel")
def reports_case_excel(case_id: int, request: Request, db: Session = Depends(get_db)):
    require_login(request, db)
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404)
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    build_case_excel(case, db, tmp.name)
    return FileResponse(tmp.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"{case.case_number}_ficha.xlsx")


@app.get("/users", response_class=HTMLResponse)
def users_page(request: Request, db: Session = Depends(get_db)):
    user = require_admin(request, db)
    users = db.query(User).order_by(User.created_at.desc()).all()
    return templates.TemplateResponse("users.html", {"request": request, "user": user, "users": users})


@app.post("/users")
def create_user(request: Request, full_name: str = Form(...), username: str = Form(...), password: str = Form(...), role: str = Form(...), db: Session = Depends(get_db)):
    require_admin(request, db)
    exists = db.query(User).filter(User.username == username).first()
    if exists:
        return RedirectResponse("/users?error=1", status_code=302)
    db.add(User(full_name=full_name, username=username, password=password, role=role))
    db.commit()
    return RedirectResponse("/users", status_code=302)



@app.get("/api/stores/{store_code}")
def get_store_data(store_code: str, request: Request, db: Session = Depends(get_db)):
    require_login(request, db)
    row = next((r for r in STORE_CATALOG if str(r.get("store_code")) == str(store_code)), None)
    if row:
        return row
    store = db.query(Store).filter(Store.store_code == store_code).first()
    if not store:
        return JSONResponse({"error": "not_found"}, status_code=404)
    return {
        "store_code": store.store_code,
        "central_code": store.central_code,
        "central_name": central_name_for(db, store.central_code),
        "format": store.format,
        "store_name": store.store_name,
        "region": store.region,
        "district": store.district,
        "zone": store.zone,
        "lat": store.lat,
        "lng": store.lng,
        "province": store.province,
        "canton": store.canton,
        "subzone": store.subzone,
        "police_district": store.police_district,
        "circuit": store.circuit,
        "subcircuit": store.subcircuit,
        "subcircuit_code": store.subcircuit_code,
    }

